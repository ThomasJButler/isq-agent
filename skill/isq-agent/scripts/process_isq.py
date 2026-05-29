"""Process a single ISQ end to end against the running rag-service.

Usage: process_isq.py <path-to-isq.pdf-or-xlsx>

Validates the input, extracts the questions, calls /process-questionnaire to answer the whole
questionnaire and assemble the canonical envelope, then writes the deliverables to ./outputs/.
The JSON deliverable is always written (stdlib only); the DOCX (and XLSX, for XLSX inputs) are
rendered when the rag-service package is importable (set ISQ_AGENT_REPO to the repo root).

Exit 0 on success, 1 on a usage or input error. Network and parsing imports are deferred into
the orchestration helpers so the argument-validation path needs nothing beyond the stdlib.
"""

import json
import os
import sys
from pathlib import Path

BASE_URL = os.environ.get("ISQ_AGENT_URL", "http://localhost:8000")
SUPPORTED = (".pdf", ".xlsx")


def main(argv: list[str] | None = None) -> int:
    argv = list(sys.argv[1:] if argv is None else argv)
    if len(argv) != 1:
        print("Usage: process_isq.py <path-to-isq.pdf-or-xlsx>")
        return 1

    input_path = Path(argv[0])
    if not input_path.exists():
        print(f"File not found: {input_path}")
        return 1

    suffix = input_path.suffix.lower()
    if suffix not in SUPPORTED:
        print(f"Unsupported file type: {suffix}. Use PDF or XLSX.")
        return 1

    return _process(input_path)


def _process(input_path: Path) -> int:
    """Extract questions, answer the whole questionnaire, and write the outputs."""
    import httpx

    source_format = "pdf" if input_path.suffix.lower() == ".pdf" else "xlsx_rows"
    extract_body: dict = {"source_format": source_format, "filename": input_path.name}
    if source_format == "pdf":
        extract_body["source_text"] = _pdf_text(input_path)
    else:
        extract_body["source_rows"] = _xlsx_rows(input_path)

    with httpx.Client(base_url=BASE_URL, timeout=120.0) as client:
        print(f"Extracting questions from {input_path.name}...")
        extracted = client.post("/extract-questions", json=extract_body)
        extracted.raise_for_status()
        questions = [
            {"question_id": q["question_id"], "text": q["text"], "index": q["index"]}
            for q in extracted.json()["questions"]
        ]

        print(f"Answering {len(questions)} question(s)...")
        processed = client.post(
            "/process-questionnaire",
            json={
                "origin": input_path.stem,
                "filename": input_path.name,
                "questions": questions,
            },
        )
        processed.raise_for_status()
        canonical = processed.json()

    return _write_outputs(input_path, canonical)


def _pdf_text(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def _xlsx_rows(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    sheet = workbook.active
    return [
        {f"col_{i}": ("" if cell is None else str(cell)) for i, cell in enumerate(row)}
        for row in sheet.iter_rows(values_only=True)
    ]


def _write_outputs(input_path: Path, canonical: dict) -> int:
    output_dir = Path("outputs")
    output_dir.mkdir(exist_ok=True)
    stem = input_path.stem

    json_path = output_dir / f"{stem}_response.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(canonical, f, indent=2, ensure_ascii=False)
        f.write("\n")
    written = [json_path]
    written.extend(_render_rich(canonical, output_dir, stem, input_path))

    meta = canonical.get("questionnaire_meta", {})
    summary = canonical.get("summary_metrics", {})
    print(
        f"Done. {meta.get('total_questions', 0)} question(s), "
        f"{summary.get('questions_flagged_for_review', 0)} flagged for review, "
        f"cost ${summary.get('total_cost_usd', 0.0):.4f}."
    )
    for path in written:
        print(f"  {path}")
    return 0


def _render_rich(
    canonical: dict, output_dir: Path, stem: str, input_path: Path
) -> list:
    """Render DOCX (and XLSX for XLSX inputs) when the rag-service renderers are importable.
    Best-effort: the JSON deliverable is always produced regardless of this step."""
    repo = os.environ.get("ISQ_AGENT_REPO")
    if repo:
        rag_service = str(Path(repo) / "rag-service")
        if rag_service not in sys.path:
            sys.path.insert(0, rag_service)
    try:
        from app.render.render_docx import render_docx
    except Exception:
        print(
            "  (DOCX/XLSX skipped: set ISQ_AGENT_REPO to the repo root to render them)"
        )
        return []

    written = [Path(render_docx(canonical, str(output_dir / f"{stem}_response.docx")))]
    if input_path.suffix.lower() == ".xlsx":
        from app.render.render_xlsx import render_xlsx

        written.append(
            Path(
                render_xlsx(
                    canonical,
                    str(output_dir / f"{stem}_response.xlsx"),
                    str(input_path),
                )
            )
        )
    return written


if __name__ == "__main__":
    sys.exit(main())
