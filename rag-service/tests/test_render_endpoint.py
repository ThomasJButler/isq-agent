"""Tests for POST /render (Plan 10 — HTTP render API for the n8n workflow).

The n8n container can't import the Python renderers the way the skill does, so it fetches
rendered deliverables over HTTP. This endpoint takes the canonical envelope (the
/process-questionnaire output) plus a target format and returns the rendered file.
"""

import io
import json
import os

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.main import app

client = TestClient(app)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _envelope(n_answers=2):
    samples = [
        (
            "q-1",
            "Do you encrypt data at rest?",
            "Yes. Northstar Labs encrypts all data at rest using AES-256.",
        ),
        (
            "q-2",
            "Where is your data stored?",
            "All data is stored in the UK (AWS eu-west-2).",
        ),
    ]
    answers = [
        {
            "question_id": qid,
            "question_text": qtext,
            "answer": ans,
            "citations": [{"source_id": "isp-s0", "text_snippet": "Northstar Labs..."}],
            "confidence": {
                "score": 0.9,
                "dimensions": {
                    "cites_policy": 0.9,
                    "on_topic": 0.95,
                    "vendor_tone": 0.9,
                    "complete": 0.9,
                },
                "needs_review": False,
                "review_reason": None,
            },
            "metrics": {
                "tokens_in": 100,
                "tokens_out": 50,
                "cost_usd": 0.002,
                "latency_ms": 1200.0,
            },
        }
        for qid, qtext, ans in samples[:n_answers]
    ]
    return {
        "questionnaire_meta": {
            "origin": "Sunflowers Charity",
            "filename": "Sunflowers.pdf",
            "received_at": "2026-05-29T10:00:00Z",
            "completed_at": "2026-05-29T10:00:42Z",
            "total_questions": n_answers,
        },
        "answers": answers,
        "summary_metrics": {
            "total_cost_usd": 0.004,
            "total_tokens": 300,
            "total_latency_ms": 2400.0,
            "questions_flagged_for_review": 0,
            "average_confidence": 0.9,
            "flagged_question_indices": [],
            "banner": None,
        },
    }


def _render(fmt, envelope=None, source_bytes=None):
    envelope = _envelope() if envelope is None else envelope
    data = {"format": fmt, "envelope": json.dumps(envelope)}
    files = (
        {"source": ("source.xlsx", source_bytes, XLSX_MIME)} if source_bytes else None
    )
    return client.post("/render", data=data, files=files)


def _source_workbook() -> bytes:
    """A minimal source questionnaire with a Response column to overlay onto."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Question", "Response"])
    ws.append(["Do you encrypt data at rest?", ""])
    ws.append(["Where is your data stored?", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_render_json_returns_canonical_envelope():
    resp = _render("json")
    assert resp.status_code == 200
    assert "application/json" in resp.headers["content-type"]
    body = resp.json()
    assert body["questionnaire_meta"]["origin"] == "Sunflowers Charity"
    assert len(body["answers"]) == 2


def test_render_docx_returns_attachment():
    resp = _render("docx")
    assert resp.status_code == 200
    assert "officedocument" in resp.headers["content-type"]
    assert resp.content[:2] == b"PK"  # a .docx is a zip container
    assert "attachment" in resp.headers.get("content-disposition", "").lower()


def test_render_xlsx_standalone_when_no_source():
    resp = _render("xlsx")
    assert resp.status_code == 200
    assert "officedocument" in resp.headers["content-type"]
    workbook = load_workbook(io.BytesIO(resp.content))
    assert "Summary" in workbook.sheetnames
    blob = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Do you encrypt data at rest?" in blob
    assert "AES-256" in blob


def test_render_xlsx_overlays_onto_source():
    resp = _render("xlsx", source_bytes=_source_workbook())
    assert resp.status_code == 200
    workbook = load_workbook(io.BytesIO(resp.content))
    sheet = workbook.active
    assert sheet.cell(row=2, column=2).value == (
        "Yes. Northstar Labs encrypts all data at rest using AES-256."
    )
    assert "Summary" in workbook.sheetnames


def test_render_rejects_unknown_format():
    resp = _render("pdf")
    assert resp.status_code in (400, 422)


def test_render_rejects_malformed_envelope():
    resp = client.post("/render", data={"format": "json", "envelope": "{not json"})
    assert resp.status_code in (400, 422)


def test_render_cleans_up_tempdir(monkeypatch):
    """Each render must not leak its temp directory once the response is served."""
    import app.api.render as render_mod

    created: list[str] = []
    real_mkdtemp = render_mod.tempfile.mkdtemp

    def tracking_mkdtemp(*args, **kwargs):
        path = real_mkdtemp(*args, **kwargs)
        created.append(path)
        return path

    monkeypatch.setattr(render_mod.tempfile, "mkdtemp", tracking_mkdtemp)
    resp = _render("json")
    assert resp.status_code == 200
    assert created, "expected the endpoint to create a temp directory"
    for path in created:
        assert not os.path.exists(path), f"temp directory leaked: {path}"


def test_render_echoes_request_id():
    """/render echoes an inbound X-Request-Id so n8n can correlate the call (like /answer)."""
    resp = client.post(
        "/render",
        data={"format": "json", "envelope": json.dumps(_envelope())},
        headers={"X-Request-Id": "isq-run-99"},
    )
    assert resp.status_code == 200
    assert resp.headers.get("x-request-id") == "isq-run-99"
