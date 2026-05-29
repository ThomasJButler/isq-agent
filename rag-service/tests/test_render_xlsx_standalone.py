"""Unit test for the standalone (no-source) XLSX render path added in Plan 10.

render_xlsx overlays onto a source workbook when one is given; for PDF inputs there is no
source, so it must build a fresh workbook from the envelope instead.
"""

from openpyxl import load_workbook

from app.render.render_xlsx import (
    XLSX_FLAGGED_FILL,
    XLSX_REVIEW_PREFIX,
    render_xlsx,
)


def _envelope():
    return {
        "questionnaire_meta": {
            "origin": "Sunflowers Charity",
            "filename": "Sunflowers.pdf",
            "received_at": "2026-05-29T10:00:00Z",
            "completed_at": "2026-05-29T10:00:42Z",
            "total_questions": 1,
        },
        "answers": [
            {
                "question_id": "q-1",
                "question_text": "Do you encrypt data at rest?",
                "answer": "Yes. Northstar Labs encrypts all data at rest using AES-256.",
                "citations": [],
                "confidence": {
                    "score": 0.9,
                    "needs_review": False,
                    "review_reason": None,
                },
                "metrics": {
                    "tokens_in": 100,
                    "tokens_out": 50,
                    "cost_usd": 0.002,
                    "latency_ms": 1200.0,
                },
            }
        ],
        "summary_metrics": {
            "total_cost_usd": 0.002,
            "total_tokens": 150,
            "total_latency_ms": 1200.0,
            "questions_flagged_for_review": 0,
            "banner": None,
        },
    }


def _flagged_envelope():
    env = _envelope()
    answer = env["answers"][0]
    answer["confidence"]["needs_review"] = True
    answer["confidence"]["review_reason"] = "Scope mismatch — needs a human."
    env["summary_metrics"]["questions_flagged_for_review"] = 1
    return env


def test_render_xlsx_standalone_without_source(tmp_path):
    out = tmp_path / "standalone.xlsx"
    result = render_xlsx(_envelope(), str(out))  # no source_path -> standalone
    assert result == str(out)

    workbook = load_workbook(str(out))
    assert "Summary" in workbook.sheetnames
    blob = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Do you encrypt data at rest?" in blob
    assert "AES-256" in blob


def test_render_xlsx_standalone_styles_flagged_answer(tmp_path):
    """A flagged answer in the standalone path gets the same [REVIEW] prefix + yellow
    fill as the overlay path, so flagged answers stand out regardless of input format."""
    out = tmp_path / "flagged.xlsx"
    render_xlsx(_flagged_envelope(), str(out))

    sheet = load_workbook(str(out))["Responses"]
    answer_cell = sheet.cell(row=2, column=3)  # row 2 = first data row, col 3 = Answer
    assert answer_cell.value.startswith(XLSX_REVIEW_PREFIX)
    assert answer_cell.fill.start_color.rgb.endswith(XLSX_FLAGGED_FILL)
