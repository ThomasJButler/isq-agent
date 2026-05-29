"""
Tests for the XLSX overlay renderer (Plan 9).

Written FIRST per TDD discipline. Implementation in app/render/render_xlsx.py
follows. Unlike DOCX (typeset from scratch), XLSX uses an OVERLAY strategy: open a
copy of the source questionnaire, find the Response column, and populate each
question's Response cell — flagged answers get a yellow fill + a [REVIEW] prefix.
A second "Summary" sheet carries the run summary (and an all-flagged banner).

Because overlay needs the original file, render_xlsx takes a third arg, source_path,
which the typeset renderers (docx/json) don't. Answers map to data rows by position:
both are 1-based ordinals from the same extraction pass, in order.
"""

import os

import pytest
from openpyxl import Workbook, load_workbook

from app.render.render_xlsx import render_xlsx


def _make_source_xlsx(path):
    """Write a source ISQ mirroring the real corpus shape (title / blank / header /
    numbered questions with empty Response cells)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Simple Salvage — Supplier Information Security Questionnaire"
    # row 2 left blank
    ws["A3"], ws["B3"], ws["C3"] = "#", "Question", "Response"
    questions = [
        "Do you enforce multi-factor authentication (MFA) on all systems?",
        "Do you maintain an incident response plan?",
        "How is privileged access to OT systems controlled?",
    ]
    for i, q in enumerate(questions, start=1):
        ws.cell(row=3 + i, column=1, value=i)
        ws.cell(row=3 + i, column=2, value=q)
    wb.save(str(path))
    return str(path)


@pytest.fixture
def source_xlsx(tmp_path):
    """A realistic source questionnaire to overlay onto."""
    return _make_source_xlsx(tmp_path / "source.xlsx")


def _canonical(answers, banner=None, flagged=0):
    """Build a canonical envelope around the given answer dicts."""
    return {
        "questionnaire_meta": {
            "origin": "Simple Salvage",
            "filename": "Simple_Salvage_Basic_ISQ.xlsx",
            "received_at": "2026-05-25T19:30:00Z",
            "completed_at": "2026-05-25T19:30:42Z",
            "total_questions": len(answers),
        },
        "answers": answers,
        "summary_metrics": {
            "total_cost_usd": 0.0072,
            "total_tokens": 2195,
            "total_latency_ms": 3020,
            "questions_flagged_for_review": flagged,
            "banner": banner,
        },
    }


def _answer(
    question_text, answer, *, needs_review=False, review_reason=None, score=0.9
):
    return {
        "question_id": "q",
        "question_text": question_text,
        "answer": answer,
        "citations": [],
        "confidence": {
            "score": score,
            "needs_review": needs_review,
            "review_reason": review_reason,
        },
        "metrics": {"cost_usd": 0.0024, "latency_ms": 1000},
    }


@pytest.fixture
def canonical_three(source_xlsx):
    """Three answers aligned 1:1 with the source data rows; the third is flagged."""
    return _canonical(
        [
            _answer(
                "Do you enforce multi-factor authentication (MFA) on all systems?",
                "Yes, MFA is mandatory on all systems.",
            ),
            _answer(
                "Do you maintain an incident response plan?", "Yes, reviewed annually."
            ),
            _answer(
                "How is privileged access to OT systems controlled?",
                "Northstar Labs does not operate OT systems.",
                needs_review=True,
                review_reason="Scope mismatch — software-only.",
                score=0.55,
            ),
        ],
        flagged=1,
    )


def _response_cells(path):
    """Return the Response-column values (col C) for the data rows of a saved sheet."""
    ws = load_workbook(str(path))["Sheet1"]
    return [ws.cell(row=r, column=3).value for r in range(4, 7)]


class TestRenderXlsx:
    def test_writes_file_to_path(self, canonical_three, source_xlsx, tmp_path):
        """render_xlsx writes a file and returns the path it wrote."""
        out = tmp_path / "out.xlsx"
        result = render_xlsx(canonical_three, str(out), source_xlsx)
        assert os.path.exists(result)
        assert result == str(out)

    def test_output_is_valid_xlsx(self, canonical_three, source_xlsx, tmp_path):
        """The written file opens cleanly as a workbook."""
        out = tmp_path / "out.xlsx"
        render_xlsx(canonical_three, str(out), source_xlsx)
        assert load_workbook(str(out)) is not None

    def test_does_not_mutate_source(self, canonical_three, source_xlsx, tmp_path):
        """The overlay writes a new file; the source Response cells stay empty."""
        out = tmp_path / "out.xlsx"
        render_xlsx(canonical_three, str(out), source_xlsx)
        assert _response_cells(source_xlsx) == [None, None, None]

    def test_populates_response_column(self, canonical_three, source_xlsx, tmp_path):
        """Every question's Response cell is filled with its answer text."""
        out = tmp_path / "out.xlsx"
        render_xlsx(canonical_three, str(out), source_xlsx)
        responses = _response_cells(out)
        assert responses[0] == "Yes, MFA is mandatory on all systems."
        assert responses[1] == "Yes, reviewed annually."

    def test_preserves_source_structure(self, canonical_three, source_xlsx, tmp_path):
        """Title row and header row survive the overlay unchanged."""
        out = tmp_path / "out.xlsx"
        render_xlsx(canonical_three, str(out), source_xlsx)
        ws = load_workbook(str(out))["Sheet1"]
        assert (
            ws["A1"].value
            == "Simple Salvage — Supplier Information Security Questionnaire"
        )
        assert (ws["A3"].value, ws["B3"].value, ws["C3"].value) == (
            "#",
            "Question",
            "Response",
        )

    def test_flagged_cell_has_review_prefix(
        self, canonical_three, source_xlsx, tmp_path
    ):
        """A flagged answer's Response cell is prefixed with [REVIEW]."""
        out = tmp_path / "out.xlsx"
        render_xlsx(canonical_three, str(out), source_xlsx)
        assert _response_cells(out)[2].startswith("[REVIEW] ")

    def test_highlights_flagged_cell_yellow(
        self, canonical_three, source_xlsx, tmp_path
    ):
        """A flagged Response cell gets the light-yellow fill (#FFEB9C)."""
        out = tmp_path / "out.xlsx"
        render_xlsx(canonical_three, str(out), source_xlsx)
        ws = load_workbook(str(out))["Sheet1"]
        fill = ws.cell(row=6, column=3).fill
        assert fill.fill_type == "solid"
        assert fill.start_color.rgb.endswith("FFEB9C")

    def test_unflagged_cell_not_highlighted(
        self, canonical_three, source_xlsx, tmp_path
    ):
        """An unflagged Response cell has no fill and no [REVIEW] prefix."""
        out = tmp_path / "out.xlsx"
        render_xlsx(canonical_three, str(out), source_xlsx)
        ws = load_workbook(str(out))["Sheet1"]
        cell = ws.cell(row=4, column=3)
        assert not str(cell.value).startswith("[REVIEW]")
        assert cell.fill.fill_type in (None, "none")

    def test_includes_summary_sheet(self, canonical_three, source_xlsx, tmp_path):
        """A second 'Summary' worksheet is appended."""
        out = tmp_path / "out.xlsx"
        render_xlsx(canonical_three, str(out), source_xlsx)
        wb = load_workbook(str(out))
        assert "Summary" in wb.sheetnames

    def test_summary_sheet_reports_flagged_count(
        self, canonical_three, source_xlsx, tmp_path
    ):
        """The Summary sheet states how many answers were flagged."""
        out = tmp_path / "out.xlsx"
        render_xlsx(canonical_three, str(out), source_xlsx)
        ws = load_workbook(str(out))["Summary"]
        text = "\n".join(
            str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
        )
        assert "Flagged for review" in text
        assert "1" in text

    def test_all_flagged_banner_in_summary(self, source_xlsx, tmp_path):
        """An all_flagged run writes the banner headline into the Summary sheet."""
        canonical = _canonical(
            [
                _answer(
                    "Do you enforce multi-factor authentication (MFA) on all systems?",
                    "n/a",
                    needs_review=True,
                    review_reason="x",
                    score=0.4,
                ),
            ],
            banner="all_flagged",
            flagged=1,
        )
        out = tmp_path / "out.xlsx"
        render_xlsx(canonical, str(out), source_xlsx)
        ws = load_workbook(str(out))["Summary"]
        text = "\n".join(
            str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
        ).upper()
        assert "ALL ANSWERS FLAGGED" in text

    def test_no_matrix_terminology_in_output(
        self, canonical_three, source_xlsx, tmp_path
    ):
        """Defence-in-depth: the renderer never propagates Matrix terms."""
        out = tmp_path / "out.xlsx"
        render_xlsx(canonical_three, str(out), source_xlsx)
        wb = load_workbook(str(out))
        text = "\n".join(
            str(c.value)
            for sheet in wb.worksheets
            for row in sheet.iter_rows()
            for c in row
            if c.value is not None
        ).lower()
        for term in [
            "morpheus",
            "the matrix",
            "neo ",
            "white rabbit",
            "redpill",
            "bluepill",
        ]:
            assert term not in text, f"Matrix term '{term}' leaked into XLSX output"

    def test_handles_empty_answers(self, source_xlsx, tmp_path):
        """An empty run still produces a valid file with a Summary sheet."""
        canonical = _canonical([])
        out = tmp_path / "out.xlsx"
        result = render_xlsx(canonical, str(out), source_xlsx)
        assert os.path.exists(result)
        assert "Summary" in load_workbook(str(out)).sheetnames
