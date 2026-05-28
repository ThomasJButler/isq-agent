"""
Document processing utilities for RAG knowledge base (Plan 4).
Handles PDF, DOCX, and XLSX extraction with error handling.
"""

from pathlib import Path
from typing import Any

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


class DocumentProcessingError(Exception):
    """Raised when document processing fails."""
    pass


def process_document(file_path: Path | str) -> dict[str, Any]:
    """
    Process a document file and extract its content.

    Supported formats:
    - PDF: Extracts text, page count, and per-page text
    - DOCX: Extracts text from paragraphs
    - XLSX: Extracts rows as structured data (for historical ISQs)

    Args:
        file_path: Path to the document file

    Returns:
        Dict with extracted content. Structure depends on file type:
        - PDF: {"text": str, "page_count": int, "pages": [{"page_number": int, "text": str}, ...]}
        - DOCX: {"text": str}
        - XLSX: {"rows": [{"question_text": str, "answer_text": str, ...}, ...]}

    Raises:
        ValueError: If file type is unsupported
        DocumentProcessingError: If processing fails
    """
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    suffix = file_path.suffix.lower()

    if suffix == ".pdf":
        return _process_pdf(file_path)
    elif suffix == ".docx":
        return _process_docx(file_path)
    elif suffix == ".xlsx":
        return _process_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")


def _process_pdf(file_path: Path) -> dict[str, Any]:
    """
    Extract text from PDF with page-level granularity.

    Args:
        file_path: Path to PDF file

    Returns:
        Dict with text, page_count, and pages list
    """
    try:
        reader = PdfReader(file_path)

        pages = []
        all_text = []

        for page_num, page in enumerate(reader.pages, start=1):
            page_text = page.extract_text()
            pages.append({
                "page_number": page_num,
                "text": page_text,
            })
            all_text.append(page_text)

        return {
            "text": "\n".join(all_text),
            "page_count": len(reader.pages),
            "pages": pages,
        }

    except Exception as e:
        raise DocumentProcessingError(f"Failed to process PDF {file_path.name}: {e}")


def _process_docx(file_path: Path) -> dict[str, Any]:
    """
    Extract text from DOCX file.

    Args:
        file_path: Path to DOCX file

    Returns:
        Dict with text key
    """
    try:
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n\n".join(paragraphs)

        return {
            "text": text,
        }

    except Exception as e:
        raise DocumentProcessingError(f"Failed to process DOCX {file_path.name}: {e}")


def _process_xlsx(file_path: Path) -> dict[str, Any]:
    """
    Extract structured data from XLSX file (for historical ISQs).

    Expects columns like: Question | Answer
    First row is treated as header.

    Args:
        file_path: Path to XLSX file

    Returns:
        Dict with rows list containing question/answer pairs
    """
    try:
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        rows = []
        header_row = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                # First row is header
                header_row = [str(cell).lower() if cell else f"col_{i}" for i, cell in enumerate(row)]
                continue

            # Skip empty rows
            if not any(cell for cell in row):
                continue

            # Build row dict
            row_dict = {}
            for col_idx, cell_value in enumerate(row):
                if col_idx < len(header_row):
                    col_name = header_row[col_idx]

                    # Normalize column names for question/answer detection
                    if "question" in col_name:
                        row_dict["question_text"] = str(cell_value) if cell_value else ""
                    elif "answer" in col_name:
                        row_dict["answer_text"] = str(cell_value) if cell_value else ""
                    else:
                        row_dict[col_name] = cell_value

            # Only include rows with question or answer data
            if row_dict.get("question_text") or row_dict.get("answer_text"):
                rows.append(row_dict)

        return {
            "rows": rows,
        }

    except Exception as e:
        raise DocumentProcessingError(f"Failed to process XLSX {file_path.name}: {e}")
