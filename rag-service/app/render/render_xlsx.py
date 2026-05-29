"""
XLSX overlay renderer (Plan 9).

Unlike the typeset DOCX renderer, this OVERLAYS the canonical answers onto a copy of
the original questionnaire: it opens the source workbook, finds the Response column,
and writes each answer into the matching row. Flagged answers get a yellow fill + a
[REVIEW] prefix so a reviewer spots them at a glance. A second "Summary" sheet carries
the run summary (and the all-flagged banner when every answer is flagged).

Overlay needs the original file, so render_xlsx takes a third arg (source_path) the
typeset renderers don't. Answers map to data rows by position — both are 1-based
ordinals from the same extraction pass, in order. The source file is never mutated:
we load it, populate in memory, and save to output_path.
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.render.shared import (
    ALL_FLAGGED_BODY,
    ALL_FLAGGED_HEADLINE,
    format_currency,
    format_duration,
)

# XLSX-only styling (Plan 9 Section 3/5). openpyxl PatternFill takes a bare hex (no #).
# Kept here, not in shared.py, since only the overlay renderer fills cells / prefixes
# in-cell text — keeping shared.py untouched also keeps the TDD commit split clean.
XLSX_FLAGGED_FILL = "FFEB9C"  # light yellow — flagged Response cells
XLSX_REVIEW_PREFIX = "[REVIEW] "  # prepended in-cell before a flagged answer

_FLAGGED_FILL = PatternFill(
    start_color=XLSX_FLAGGED_FILL, end_color=XLSX_FLAGGED_FILL, fill_type="solid"
)


def render_xlsx(canonical: dict, output_path: str, source_path: str) -> str:
    """Overlay the canonical answers onto a copy of the source questionnaire.

    Returns the output path. The source file at source_path is not modified.
    """
    answers = canonical.get("answers", [])
    summary = canonical.get("summary_metrics", {})
    meta = canonical.get("questionnaire_meta", {})

    workbook = load_workbook(source_path)
    sheet = workbook.active
    _overlay_answers(sheet, answers)
    _add_summary_sheet(workbook, meta, summary)

    workbook.save(output_path)
    return output_path


def _find_response_column(sheet: Worksheet) -> tuple[int, int] | None:
    """Locate the (header_row, response_col) by scanning for a 'Response' header.

    Returns None if no Response column is found (e.g. a malformed source) so the
    caller can skip overlay rather than crash.
    """
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().lower() == "response":
                return cell.row, cell.column
    return None


def _overlay_answers(sheet: Worksheet, answers: list[dict]) -> None:
    """Write each answer into its Response cell, in source row order."""
    location = _find_response_column(sheet)
    if location is None or not answers:
        return
    header_row, response_col = location

    # Data rows are the rows after the header that carry a question (col left of
    # Response is non-empty). Match answers to these rows by position.
    question_col = max(1, response_col - 1)
    data_rows = [
        r
        for r in range(header_row + 1, sheet.max_row + 1)
        if sheet.cell(row=r, column=question_col).value not in (None, "")
    ]

    for answer, row in zip(answers, data_rows):
        confidence = answer.get("confidence") or {}
        text = answer.get("answer", "")
        cell = sheet.cell(row=row, column=response_col)
        if confidence.get("needs_review"):
            cell.value = f"{XLSX_REVIEW_PREFIX}{text}"
            cell.fill = _FLAGGED_FILL
        else:
            cell.value = text


def _add_summary_sheet(workbook, meta: dict, summary: dict) -> None:
    """Append a 'Summary' worksheet with the run totals + optional banner."""
    sheet = workbook.create_sheet("Summary")
    row = 1

    if summary.get("banner") == "all_flagged":
        sheet.cell(row=row, column=1, value=ALL_FLAGGED_HEADLINE).fill = _FLAGGED_FILL
        sheet.cell(row=row + 1, column=1, value=ALL_FLAGGED_BODY)
        row += 3

    pairs = [
        ("Total questions:", meta.get("total_questions", 0)),
        ("Flagged for review:", summary.get("questions_flagged_for_review", 0)),
        ("Total cost:", format_currency(summary.get("total_cost_usd", 0.0))),
        ("Total time:", format_duration(summary.get("total_latency_ms", 0))),
    ]
    for label, value in pairs:
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=value)
        row += 1
